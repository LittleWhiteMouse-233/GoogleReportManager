import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from .Base import Base
from .SuiteReport import SuiteReport
from . import utils
from . import workbookProcess as wbP


class XTSReport(Base):
    def __init__(self, xts_dir: str):
        super().__init__()
        self.logger.info("XTSReport loading start.")
        self.__xts_path = utils.absolute_path(xts_dir)
        valid_paths = self.search_valid_suite_report(self.__xts_path)
        if len(valid_paths) == 0:
            self.logger.warning("No valid suite report path.")
        self.__xts_report = self.__load_xts_report(valid_paths)
        self.logger.info("XTSReport loading completed.")
        self.__verify_data()

    @classmethod
    def search_valid_suite_report(cls, xts_path: str):
        path_list = utils.dirs_sort_by_create(xts_path, reverse=True)
        valid_paths = [path for path in path_list if SuiteReport.is_suite_report(path)]
        cls.logger.debug("Valid suite report paths [%d]: \n" + '\n'.join(valid_paths), len(valid_paths))
        return valid_paths

    def __verify_data(self):
        def func_assert():
            assert len(self.__xts_report) > 0

        return self.try_except(func_assert)

    @classmethod
    def __load_xts_report(cls, path_list: list[str]):
        cls.logger.info("Loading xts report.")
        xts_report = dict()
        for path in path_list:
            suite_report = SuiteReport(path)
            suite_name = suite_report.get_suite_name()
            if suite_name not in xts_report.keys():
                xts_report[suite_name] = suite_report
        sorted_xts = dict(sorted(xts_report.items(), key=lambda x: x[0], reverse=False))
        cls.logger.info("Loading xts report completed.")
        return sorted_xts

    def __gen_device_info_sheet(self):
        self.logger.info("Generating device info sheet.")
        device_info_list = []
        for report in self.__xts_report.values():
            device_info = report.get_device_info_table()
            if device_info is not None:
                device_info_list.append(device_info)
        if not device_info_list:
            self.logger.warning("Unable to generate device info sheet because not valid device info table.")
            return None
        sheet = device_info_list[0]
        column = sheet.columns[0]
        for di in device_info_list[1:]:
            sheet = pd.merge(sheet, di, how='outer', on=column)
        device_info_sheet = wbP.b_header(sheet)
        self.logger.info("Generated device info sheet.")
        return device_info_sheet

    def __gen_xts_abstract_table(self):
        self.logger.info("Generating XTS abstract table.")
        if not self.__xts_report:
            self.logger.error("Unable to generate XTS abstract table because not valid suite report.")
            raise ValueError
        abstract_list = [report.get_abstract_table() for report in self.__xts_report.values()]
        xts_abstract = wbP.b_header(pd.concat(abstract_list).reset_index(drop=True))
        self.logger.info("Generated XTS abstract table.")
        return xts_abstract

    def __gen_top_sheet(self):
        self.logger.info("Generating top sheet.")
        if 'CTS' not in self.__xts_report.keys():
            self.logger.error("Unable to generate XTS top table because not suite report of CTS.")
            raise ValueError
        top_table = self.__xts_report['CTS'].get_top_table()
        xts_abstract = self.__gen_xts_abstract_table()
        top_sheet = pd.concat([top_table, xts_abstract]).reset_index(drop=True)
        self.logger.info("Generated top sheet.")
        return top_sheet

    def get_workbook(self):
        self.logger.info("Getting workbook.")
        wb = Workbook()
        top_sheet = self.__gen_top_sheet()
        ws_sum: Worksheet = wb.active
        ws_sum.title = 'SUM'
        self.logger.info("Entering [%s] sheet and render styles" % ws_sum.title)
        ws_sum = self.enter_sheet(ws_sum, top_sheet,
                                  column_width={
                                      'A': 20,
                                      'B': 25,
                                      'C': 10,
                                      'D': 10,
                                      'E': 20,
                                      'F': 10,
                                      'G': 100,
                                  })
        wb.worksheets[0] = ws_sum
        device_info_sheet = self.__gen_device_info_sheet()
        if device_info_sheet is not None:
            ws_di: Worksheet = wb.create_sheet('device info')
            column_width = {}
            column_width.update(map(lambda x: (x, 50), range(1, device_info_sheet.shape[1] + 1)))
            self.logger.info("Entering [%s] sheet and render styles" % ws_di.title)
            ws_di = self.enter_sheet(ws_di, device_info_sheet, column_width=column_width)
            wb.worksheets[1] = ws_di
        for suite, report in self.__xts_report.items():
            ws_suite = wb.create_sheet(suite)
            self.logger.info("Entering [%s] sheet and render styles" % ws_suite.title)
            ws_suite = self.enter_sheet(ws_suite, report.get_suite_sheet(),
                                        column_width={
                                            'A': 30,
                                            'B': 70,
                                        })
            wb.worksheets[-1] = ws_suite
        self.logger.info("Got workbook.")
        return wb

    @staticmethod
    def enter_sheet(ws: Worksheet, df: pd.DataFrame, column_width: dict[str, int] = None):
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        wsp = wbP.WorksheetProcessor(ws)
        wsp.set_column_width(column_width)
        wsp.render_cell_styles()
        return wsp.get_worksheet()
