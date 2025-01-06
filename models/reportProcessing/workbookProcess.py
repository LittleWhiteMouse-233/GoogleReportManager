from enum import Enum
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Font
from copy import copy
import re
from typing import Callable
from . import utils


class WorksheetProcessor:
    def __init__(self, ws: Worksheet):
        self.__ws = ws
        self.__max_row = ws.max_row
        self.__max_col = ws.max_column
        self.__range_row = range(1, self.__max_row + 1)
        self.__range_col = range(1, self.__max_col + 1)

    def get_worksheet(self):
        return self.__ws

    @staticmethod
    def alpha_to_decimal_num(alpha_str: str):
        if not alpha_str.isalpha():
            raise ValueError("The parameter \'alpha_str\' has non-alphabetic characters: " + str(alpha_str))
        num_list = list(map(lambda c: ord(c.upper()) - ord('A') + 1, list(alpha_str)))
        return utils.any_scale_to_decimal(num_list, 26)

    @staticmethod
    def decimal_num_to_alpha(decimal_num: int):
        if decimal_num <= 0:
            raise ValueError("The parameter \'decimal_num\'(" + str(decimal_num) + ") cannot be less than 0.")
        alpha_str = ''
        while True:
            decimal_num, mod = divmod(decimal_num, 26)
            alpha_str = chr(mod + ord('A') - 1) + alpha_str
            if decimal_num == 0:
                break
        return alpha_str

    def __activate_col(self, col: int | str):
        if type(col) is not int:
            if type(col) is not str:
                raise ValueError("Invalid type of col: %s, excepted int or str" % str(type(col)))
            col = self.alpha_to_decimal_num(col)
        if not 1 <= col <= self.__max_col:
            raise ValueError("The col(%d) is out of range(%d:%d)" % (col, 1, self.__max_col))
        return col

    def __activate_col_list(self, col_list: list[int | str]):
        if not col_list:
            return []
        return sorted(list(set(map(lambda c: self.__activate_col(c), col_list))))

    def set_column_width(self, column_width: dict[str, int]):
        for col, width in column_width.items():
            if not isinstance(width, (int, float)):
                raise ValueError("Invalid width type: %s." % type(width))
            if width <= 0:
                raise ValueError("The column width should be positive, but given: %s" % width)
            activated_col = self.__activate_col(col)
            self.__ws.column_dimensions[self.decimal_num_to_alpha(activated_col)].width = width

    def render_cell_styles(self):
        for i in self.__range_row:
            for j in self.__range_col:
                content = str(self.__ws.cell(i, j).value)
                if content != 'nan':
                    self.setting_cell_border(i, j)
                for s in list(Style):
                    s: Style
                    if re.search(s.tag, content):
                        self.__ws.cell(i, j).value = re.sub(s.tag, '', content)
                        s.func(self, i, j, **s.kwargs)

    def setting_cell_border(self, row, col, border_style='thin'):
        excepted_style = ["dashDot", "dashDotDot", "dashed", "dotted", "double", "hair", "medium", "mediumDashDot",
                          "mediumDashDotDot", "mediumDashed", "slantDashDot", "thick", "thin", "none"]
        if border_style not in excepted_style:
            raise ValueError("Invalid border_style: %s, excepted input: %s" % (border_style, excepted_style))
        self.__ws.cell(row, col).border = Border(left=Side(border_style=border_style, color='FF000000'),
                                                 right=Side(border_style=border_style, color='FF000000'),
                                                 top=Side(border_style=border_style, color='FF000000'),
                                                 bottom=Side(border_style=border_style, color='FF000000'))

    def setting_basic_font(self, row, col, name: str = None, size: int = None, bold: bool = None, color: str = None,
                           italic: bool = None, strike: bool = None):
        font_new: Font = copy(self.__ws.cell(row, col).font)
        if name:
            font_new.name = name
        if size:
            font_new.size = size
        if bold:
            font_new.bold = bold
        if color:
            font_new.color = color
        if italic:
            font_new.italic = italic
        if strike:
            font_new.strike = strike
        self.__ws.cell(row, col).font = font_new

    def setting_word_wrap(self, row, col):
        align = self.__ws.cell(row, col).alignment
        if not align.wrapText:
            align_new = copy(align)
            align_new.wrapText = True
            self.__ws.cell(row, col).alignment = align_new


class Style(Enum):
    B = '<b>', WorksheetProcessor.setting_basic_font, {'bold': True}
    WW = '<ww>', WorksheetProcessor.setting_word_wrap, {}

    def __init__(self, tag: str, func: Callable, kwargs: dict):
        self.tag = tag
        self.func = func
        self.kwargs = kwargs


def add_tag(content: str, style: Style):
    return style.tag + str(content)


def b_header(table: pd.DataFrame, reset_column=True, reset_index=True):
    column = list(map(lambda x: add_tag(x, Style.B), table.columns))
    table.columns = column
    if reset_column:
        table = utils.reset_column(table)
    if reset_index:
        table = table.reset_index(drop=True)
    return table
