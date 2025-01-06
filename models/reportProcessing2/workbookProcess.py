from enum import Enum
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import re
from typing import Callable
from . import utils


class WorksheetProcessor:
    def __init__(self, ws: Worksheet):
        self.__ws = ws
        self.__max_row = ws.max_row
        self.__max_col = ws.max_column
        self.__range_row = range(1, self.__max_row + 1)
        self.__range_col = range(1, self.__max_col + 1)

    def get_worksheet(self):
        return self.__ws

    @staticmethod
    def __alpha_to_decimal_num(alpha_str: str):
        if not alpha_str.isalpha():
            raise ValueError("The parameter \'alpha_str\' has non-alphabetic characters: " + str(alpha_str))
        ord_list = list(map(lambda c: ord(c.upper()) - ord('A') + 1, list(alpha_str)))

        def any_scale_to_decimal(num_list: list[int], origin_scale: int, natural=True):
            if natural:
                num_list.reverse()
            decimal_num = 0
            for i, num in enumerate(num_list):
                decimal_num += num * pow(origin_scale, i)
            return decimal_num

        return any_scale_to_decimal(ord_list, 26)

    @staticmethod
    def __decimal_num_to_alpha(decimal_num: int):
        if decimal_num <= 0:
            raise ValueError("The parameter \'decimal_num\'(" + str(decimal_num) + ") cannot be less than 0.")
        alpha_str = ''
        while True:
            decimal_num, mod = divmod(decimal_num, 26)
            alpha_str = chr(mod + ord('A') - 1) + alpha_str
            if decimal_num == 0:
                break
        return alpha_str

    def __activate_col(self, col: int | str):
        if type(col) is not int:
            if type(col) is not str:
                raise ValueError("Invalid type of col: %s, excepted int or str" % str(type(col)))
            col = self.__alpha_to_decimal_num(col)
        if not 1 <= col <= self.__max_col:
            raise ValueError("The col(%d) is out of range(%d:%d)" % (col, 1, self.__max_col))
        return col

    def __activate_col_list(self, col_list: list[int | str]):
        if not col_list:
            return []
        return sorted(list(set(map(lambda c: self.__activate_col(c), col_list))))

    def set_column_width(self, column_width: dict[str | int, int]):
        for col, width in column_width.items():
            if not isinstance(width, (int, float)):
                raise ValueError("Invalid width type: %s." % type(width))
            if width <= 0:
                raise ValueError("The column width should be positive, but given: %s" % width)
            activated_col = self.__activate_col(col)
            self.__ws.column_dimensions[self.__decimal_num_to_alpha(activated_col)].width = width

    def render_cell_styles(self):
        for i in self.__range_row:
            for j in self.__range_col:
                content = str(self.__ws.cell(i, j).value)
                if content != 'nan':
                    self.setting_cell_border(i, j)
                for s in list(Style):
                    s: Style
                    if re.search(s.tag, content):
                        s.func(self, i, j, **s.kwargs)
                        self.__ws.cell(i, j).value = re.sub(s.tag, '', content)

    def setting_cell_border(self, row, col, border_style='thin'):
        excepted_style = ["dashDot", "dashDotDot", "dashed", "dotted", "double", "hair", "medium", "mediumDashDot",
                          "mediumDashDotDot", "mediumDashed", "slantDashDot", "thick", "thin", "none"]
        if border_style not in excepted_style:
            raise ValueError("Invalid border_style: %s, excepted input: %s" % (border_style, excepted_style))
        self.__ws.cell(row, col).border = Border(left=Side(border_style=border_style, color='FF000000'),
                                                 right=Side(border_style=border_style, color='FF000000'),
                                                 top=Side(border_style=border_style, color='FF000000'),
                                                 bottom=Side(border_style=border_style, color='FF000000'))

    def setting_basic_font(self, row, col, name: str = None, size: int = None, bold: bool = None, color: str = None,
                           italic: bool = None, strike: bool = None):
        font_new = copy(self.__ws.cell(row, col).font)
        if name:
            font_new.name = name
        if size:
            font_new.size = size
        if bold:
            font_new.bold = bold
        if color:
            font_new.color = color
        if italic:
            font_new.italic = italic
        if strike:
            font_new.strike = strike
        self.__ws.cell(row, col).font = font_new

    def setting_word_wrap(self, row, col):
        align = self.__ws.cell(row, col).alignment
        if not align.wrapText:
            align_new = copy(align)
            align_new.wrapText = True
            self.__ws.cell(row, col).alignment = align_new

    def setting_fill_color(self, row, col, re_pattern: re.Pattern = re.compile(r'<0[xX][a-zA-Z0-9]{6}>'),
                           fill_type='solid'):
        excepted_type = ["solid", "darkDown", "darkGray", "darkGrid", "darkHorizontal", "darkTrellis", "darkUp",
                         "darkVertical", "gray0625", "gray125", "lightDown", "lightGray", "lightGrid",
                         "lightHorizontal", "lightTrellis", "lightUp", "lightVertical", "mediumGray", "none"]
        if fill_type not in excepted_type:
            raise ValueError("Invalid fill_type: %s, excepted input: %s" % (fill_type, excepted_type))
        type_index = excepted_type.index(fill_type)
        content = self.__ws.cell(row, col).value
        color = re_pattern.search(content)
        if color:
            self.__ws.cell(row, col).fill = PatternFill(patternType=excepted_type[type_index],
                                                        fgColor=color.group()[3:-1])
            self.__ws.cell(row, col).value = re_pattern.sub('', content)


class Style(Enum):
    B = '<b>', WorksheetProcessor.setting_basic_font, {'bold': True}
    WW = '<ww>', WorksheetProcessor.setting_word_wrap, {}
    RED = '<0xFF0000>', WorksheetProcessor.setting_fill_color, {}
    YELLOW = '<0xFFFF00>', WorksheetProcessor.setting_fill_color, {}
    GREEN = '<0x00FF00>', WorksheetProcessor.setting_fill_color, {}
    GRAY = '<0xAAAAAA>', WorksheetProcessor.setting_fill_color, {}
    WRONG = RED
    WARNING = YELLOW

    def __init__(self, tag: str, func: Callable, kwargs: dict):
        self.tag = tag
        self.func = func
        self.kwargs = kwargs


def add_tag(content: str, style: Style | str):
    if type(style) is Style:
        return style.tag + str(content)
    else:
        return str(style) + str(content)


def b_header(table: pd.DataFrame, reset_col=True, reset_row=True):
    column = list(map(lambda x: add_tag(x, Style.B), table.columns))
    table.columns = column
    if reset_col:
        table = utils.reset_column(table)
    if reset_row:
        table = table.reset_index(drop=True)
    return table


def highlight_map_table(map_table: pd.DataFrame, keyword: str, tag: Style, col0=0, col1=1):
    row_i = utils.locate_map_table(map_table, keyword, col=col0)
    map_table.iloc[row_i, col1] = add_tag(map_table.iloc[row_i, col1], tag)
    return map_table


def adaptive_column_width(sheet_table: pd.DataFrame):
    column_width = {}
    for i in range(sheet_table.shape[0]):
        for j in range(sheet_table.shape[1]):
            col_id = j + 1
            col_len = len(sheet_table.iloc[i, j]) + 2
            if col_id not in column_width.keys() or col_len > column_width[col_id]:
                column_width[col_id] = col_len
    return column_width


def enter_sheet(ws: Worksheet, df: pd.DataFrame, column_width: dict[str | int, int] = None):
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    wsp = WorksheetProcessor(ws)
    wsp.set_column_width(column_width)
    wsp.render_cell_styles()
    return wsp.get_worksheet()
