import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import time
from .Base import Base
from .Report import Report
from .SuiteReport import SuiteReport, NoReportException
from . import utils
from . import workbookProcess as wbP


class XTSReport(Base):
    __test_suite = [
        'CTS',
        'CTS_ON_GSI',
        'CTS_VERIFIER',
        'GTS',
        'STS',
        'TVTS',
        'VTS',
    ]

    def __init__(self, xts_dir: str, flag_unpack=False):
        super().__init__()
        self.__xts_path = utils.absolute_path(xts_dir)
        dir_path_list = utils.dirs_sort_by_create(self.__xts_path, reverse=True)
        self.__xts_report = self.__load_xts_report(dir_path_list, flag_unpack=flag_unpack)
        if len(self.__xts_report) == 0:
            raise ValueError("No valid SuiteReport in XTSReport.")
        self.__verify()
        self.__inspect_spl()
        self.__inspect_suite_summary()

    @staticmethod
    def __load_xts_report(path_list: list[str], flag_unpack=False):
        xts_report = dict()
        for path in path_list:
            try:
                suite_report = SuiteReport(path, flag_unpack=flag_unpack)
            except NoReportException:
                continue
            suite_name = suite_report.suite_name
            if suite_name not in xts_report.keys():
                xts_report[suite_name] = suite_report
        sorted_xts = dict(sorted(xts_report.items(), key=lambda x: x[0], reverse=False))
        return sorted_xts

    def __verify(self):
        assert len(self.__xts_report) > 0
        assert 'CTS' in self.__xts_report.keys()

    @property
    def xts_dir(self):
        return utils.path_basename(self.__xts_path)

    @property
    def existed_suite(self):
        return list(self.__xts_report.keys())

    def __inspect_spl(self):
        if 'CTS' not in self.__xts_report.keys():
            return
        time_format = '%Y-%m-%d'
        cts_spl_time = time.strptime(self.__xts_report['CTS'].main_summary_spl, time_format)
        if 'CTS_ON_GSI' in self.__xts_report.keys():
            gsi = self.__xts_report['CTS_ON_GSI']
            if time.strptime(gsi.main_summary_spl, time_format).tm_mon < cts_spl_time.tm_mon:
                gsi.wrong_spl = True
        if 'VTS' in self.__xts_report.keys():
            vts = self.__xts_report['VTS']
            if time.strptime(vts.main_summary_spl, time_format).tm_mon < cts_spl_time.tm_mon:
                vts.wrong_spl = True

    def __inspect_suite_summary(self):
        check_list = ['Fingerprint', 'Release (SDK)', 'ABIs']
        inconsistent = []
        for row_title in check_list:
            value_set = set()
            for suite_report in self.__xts_report.values():
                value_set.add(suite_report.search_main_summary(row_title))
            if len(value_set) != 1:
                inconsistent.append(row_title)
        for suite_report in self.__xts_report.values():
            suite_report.inconsistent_summary = inconsistent

    def __create_device_info_sheet(self):
        device_info_list = []
        for report in self.__xts_report.values():
            device_info = report.get_device_info_table()
            if device_info is not None:
                device_info_list.append(device_info)
        if not device_info_list:
            self.logger.warning("Unable to create device info sheet because not valid device info table.")
            return None
        sheet = device_info_list[0]
        on_column = sheet.columns[0]
        for di in device_info_list[1:]:
            sheet = pd.merge(sheet, di, how='outer', on=on_column)
        sheet = Report.filter_property(sheet, on=on_column)
        device_info_sheet = wbP.b_header(sheet)
        return device_info_sheet

    def __create_top_sheet(self, suite_name: str = 'CTS'):
        if suite_name not in self.__xts_report.keys():
            raise KeyError("Report of [%s] is required to create top sheet." % suite_name)
        top_table = self.__xts_report['CTS'].get_top_table()
        abstract_list = []
        pass_idx = []
        for sn in self.__test_suite:
            if sn in self.existed_suite:
                abstract_list.append(self.__xts_report[sn].get_abstract_table())
                pass_idx.append(self.existed_suite.index(sn))
            else:
                abstract_list.append(SuiteReport.get_empty_abstract_table(sn))
        for i, report in enumerate(self.__xts_report.values()):
            if i in pass_idx:
                continue
            abstract_list.append(report.get_abstract_table())
        xts_abstract = wbP.b_header(pd.concat(abstract_list).reset_index(drop=True))
        top_sheet = pd.concat([top_table, xts_abstract]).reset_index(drop=True)
        return top_sheet

    def get_workbook(self):
        wb = Workbook()
        top_sheet = self.__create_top_sheet()
        ws_sum: Worksheet = wb.active
        ws_sum.title = 'SUM'
        ws_sum = wbP.enter_sheet(ws_sum, top_sheet,
                                 column_width={
                                     'A': 20,
                                     'B': 25,
                                     'C': 10,
                                     'D': 10,
                                     'E': 20,
                                     'F': 15,
                                     'G': 100,
                                 })
        wb.worksheets[0] = ws_sum
        device_info_sheet = self.__create_device_info_sheet()
        if device_info_sheet is not None:
            ws_di: Worksheet = wb.create_sheet('device info')
            column_width = wbP.adaptive_column_width(device_info_sheet)
            ws_di = wbP.enter_sheet(ws_di, device_info_sheet, column_width=column_width)
            wb.worksheets[1] = ws_di
        for suite_name, report in self.__xts_report.items():
            ws_suite = wb.create_sheet(suite_name)
            ws_suite = wbP.enter_sheet(ws_suite, report.get_suite_sheet(),
                                       column_width={
                                           'A': 30,
                                           'B': 70,
                                       })
            wb.worksheets[-1] = ws_suite
        return wb
